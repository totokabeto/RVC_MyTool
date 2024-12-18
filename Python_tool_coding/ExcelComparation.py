#!/usr/bin/python
# -*- coding: utf-8 -*-

################################################################################
### Module   = ExcelComaprasion.py                                 ###
### Author   = TuNguyen                                                      ###
### Date     = 21/01/2021                                                    ###
###                     Revision History                                     ###
###                                                                          ###
###  1.0.0: 21/01/2021  : Initial version.                                   ###
################################################################################
###                       Import section                                     ###
################################################################################

from openpyxl import Workbook
from openpyxl import worksheet
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
import xlrd
from copy import deepcopy, copy
import sys, os
import pyexcel as p
import pyexcel_io.writers
import pyexcel_xls
import pyexcel_xlsx



def main():

    ExcelFile1 = sys.argv[1]
    ExcelFile2 = sys.argv[2]
    if ExcelFile1.split(".")[1] == "xls":
        p.save_book_as(file_name=ExcelFile1,
               dest_file_name='newFile1.xlsx')
        p.save_book_as(file_name=ExcelFile2,
               dest_file_name='newFile2.xlsx')
        ExcelFile1 = 'newFile1.xlsx'
        ExcelFile2 = 'newFile2.xlsx'

    WB1 = load_workbook(ExcelFile1, data_only=True)
    WB2 = load_workbook(ExcelFile2, data_only=True)
    listOfDiffSheets = []
    wb = Workbook()
    for sheet in WB1.worksheets:
        isDifferent = False
        newtitle = sheet.title
        if '_r' in sheet.title:
            newtitle = sheet.title.rpartition('_r')[0]
        ws = wb.create_sheet(newtitle)
        ws1 = WB1[sheet.title]
        try:
            match = False
            for sheet2 in WB2.worksheets:
                if newtitle == sheet2.title:
                    match = True
                    ws2 = WB2[sheet2.title]
            if match is False:
                continue
        except:
            print("ERROR: Can not find sheet: \"" + str(sheet.title) + "\" in \"" +str(ExcelFile2) + "\"")
            continue
        maxCol = ws1.max_column
        maxRow = ws1.max_row
        j=0
        for row in range(1, maxRow + 1):
            i=0
            if j==50:
                break
            for col in range(1, maxCol + 1):
                
                cell1 = ws1.cell(row = row, column = col)
                cell2 = ws2.cell(row = row, column = col)
                if cell1.value==None:
                    i=i+1
                if i==maxCol:
                    j=j+1
                               
                if cell1.value != cell2.value:
                    cell1.fill = PatternFill("solid", fgColor="FF0000")
                    cell2.fill = PatternFill("solid", fgColor="FF0000")
                    isDifferent = True
                leftCell = ws.cell(row = row, column = col)
                rightCell = ws.cell(row = row, column = col + maxCol + 2)
                leftCell.value = cell1.value
                rightCell.value = cell2.value
                if cell1.has_style:
                    leftCell.font = copy(cell1.font)
                    leftCell.border = copy(cell1.border)
                    leftCell.fill = copy(cell1.fill)
                    leftCell.number_format = copy(cell1.number_format)
                    leftCell.protection = copy(cell1.protection)
                    leftCell.alignment = copy(Alignment(wrap_text=False))
                if cell2.has_style:
                    rightCell.font = copy(cell2.font)
                    rightCell.border = copy(cell2.border)
                    rightCell.fill = copy(cell2.fill)
                    rightCell.number_format = copy(cell2.number_format)
                    rightCell.protection = copy(cell2.protection)
                    rightCell.alignment = copy(Alignment(wrap_text=False))
                   
            middleCell = ws.cell(row = row, column = maxCol + 2)
            middleCell.fill = PatternFill("solid", fgColor="000000")
        if isDifferent == True:
            print(" [+] " + newtitle + " : NG <<==")
            listOfDiffSheets.append(newtitle)
            ws.sheet_properties.tabColor = 'FF0000'
        else:
            print(" [+] " + newtitle + " : OK")
    print("Finish!")
    if len(listOfDiffSheets) != 0:
        print("list of different sheets: " + str(listOfDiffSheets))
    else:
        print(" - - Finish with no differences - - ")
    try:
        os.remove('newFile1.xlsx')
        os.remove('newFile2.xlsx')
    except:
        pass
    wb.save('output.xlsx')
if __name__ == "__main__":
    main()

# -----------------------------------------------------------
# end of script
# -----------------------------------------------------------